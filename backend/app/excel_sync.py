"""Utilities to synchronize Empresas/Filiais with Planilha Mestre.xlsx."""
from __future__ import annotations

import logging
import os
import re
from contextlib import contextmanager
from pathlib import Path
from dataclasses import dataclass
from typing import Dict, Iterator, List, Optional, Set

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)

EXCEL_PATH_ENV = "EXCEL_MASTER_PATH"
EMPRESAS_SHEET_ENV = "EXCEL_SHEET_EMPRESAS"
FILIAIS_SHEET_ENV = "EXCEL_SHEET_FILIAIS"

DEFAULT_EMPRESAS_SHEET = "Empresas"
DEFAULT_FILIAIS_SHEET = "Filiais"


class ExcelSyncError(RuntimeError):
    """Base class for spreadsheet sync errors."""


class ExcelSyncConfigError(ExcelSyncError):
    """Raised when EXCEL_MASTER_PATH is not configured."""


class ExcelSyncLockedError(ExcelSyncError):
    """Raised when the spreadsheet is opened/locked in Excel."""


# ---------------------------------------------------------------------------
# helpers
# ---------------------------------------------------------------------------

def _excel_path() -> Path:
    path_str = os.getenv(EXCEL_PATH_ENV, "").strip()
    if not path_str:
        raise ExcelSyncConfigError(
            "Variavel EXCEL_MASTER_PATH nao configurada."
        )
    path = Path(path_str)
    if not path.exists():
        raise ExcelSyncError(f"Planilha nao encontrada: {path}")
    return path


@contextmanager
def _open_workbook() -> Iterator[tuple[Path, "Workbook"]]:
    from openpyxl.workbook import Workbook

    path = _excel_path()
    try:
        wb = load_workbook(path)
    except PermissionError as exc:
        raise ExcelSyncLockedError(
            "Nao foi possivel acessar a planilha. Feche o arquivo no Excel e tente novamente."
        ) from exc
    except OSError as exc:
        raise ExcelSyncError(f"Falha ao abrir a planilha: {exc}") from exc

    try:
        yield path, wb
        try:
            wb.save(path)
        except PermissionError as exc:
            raise ExcelSyncLockedError(
                "Nao foi possivel salvar a planilha. Feche o arquivo no Excel e tente novamente."
            ) from exc
        except OSError as exc:
            raise ExcelSyncError(f"Falha ao salvar a planilha: {exc}") from exc
    finally:
        try:
            wb.close()
        except Exception:
            pass


def _open_workbook_values() -> tuple[Path, "Workbook"]:
    path = _excel_path()
    try:
        wb = load_workbook(path, data_only=True)
    except PermissionError as exc:
        raise ExcelSyncLockedError(
            "Nao foi possivel acessar a planilha. Feche o arquivo no Excel e tente novamente."
        ) from exc
    except OSError as exc:
        raise ExcelSyncError(f"Falha ao abrir a planilha: {exc}") from exc
    return path, wb


def _normalize_id(value: Optional[object], *, width: int) -> str:
    if value is None:
        return "".zfill(width)
    try:
        text = str(value).strip()
    except Exception:
        text = str(value)
    if not text:
        return "".zfill(width)
    if text.endswith(".0") and text[:-2].isdigit():
        text = text[:-2]
    digits = "".join(ch for ch in text if ch.isdigit())
    if digits:
        text = digits
    try:
        number = int(text)
        return f"{number:0{width}d}"
    except ValueError:
        return text.zfill(width)


def _collect_ids(sheet_name: str, *, column: int, width: int) -> Set[str]:
    path, wb = _open_workbook_values()
    try:
        if sheet_name not in wb.sheetnames:
            raise ExcelSyncError(f"Planilha '{path}' nao possui aba '{sheet_name}'")
        sheet = wb[sheet_name]
        ids: Set[str] = set()
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row or len(row) <= column:
                continue
            normalized = _normalize_id(row[column], width=width)
            if normalized.strip("0"):
                ids.add(normalized)
        return ids
    finally:
        try:
            wb.close()
        except Exception:
            pass


def _sheet_name(env_name: str, default: str) -> str:
    return os.getenv(env_name, default)


def _empresa_exists(sheet: Worksheet, id_empresa: str) -> bool:
    normalized_id = _normalize_id(id_empresa, width=4)
    existing = _collect_ids(sheet.title, column=2, width=4)
    return normalized_id in existing


def _filial_exists(sheet: Worksheet, id_empresa: str, id_unidade: str) -> bool:
    normalized_emp = _normalize_id(id_empresa, width=4)
    normalized_unit = _normalize_id(id_unidade, width=3)
    path, wb = _open_workbook_values()
    try:
        sheet_values = wb[sheet.title]
        for row in sheet_values.iter_rows(min_row=2, values_only=True):
            if not row:
                continue
            emp_id = _normalize_id(row[1] if len(row) > 1 else None, width=4)
            unit_id = _normalize_id(row[2] if len(row) > 2 else None, width=3)
            if emp_id == normalized_emp and unit_id == normalized_unit:
                return True
        return False
    finally:
        try:
            wb.close()
        except Exception:
            pass


def _path_str(base_dir: Optional[Path], nome_empresa: str, id_empresa: str, nome_unidade: str, id_unidade: str) -> str:
    if base_dir is None:
        return ""
    try:
        base = Path(base_dir)
        full = base / f"{nome_empresa} - {id_empresa}" / f"{nome_unidade} - {id_unidade}"
        return os.path.normpath(str(full))
    except Exception as exc:
        logger.warning("Falha ao montar PATH: %s", exc)
        return ""


# ---------------------------------------------------------------------------
# public API
# ---------------------------------------------------------------------------

def append_empresa(nome: str, id_empresa: str) -> bool:
    sheet_name = _sheet_name(EMPRESAS_SHEET_ENV, DEFAULT_EMPRESAS_SHEET)
    try:
        with _open_workbook() as (path, workbook):
            if sheet_name not in workbook.sheetnames:
                raise ExcelSyncError(f"Planilha '{path}' nao possui aba '{sheet_name}'")
            sheet = workbook[sheet_name]
            normalized_id = _normalize_id(id_empresa, width=4)
            if _empresa_exists(sheet, normalized_id):
                return False
            sheet.append([nome, None, normalized_id])
            return True
    except ExcelSyncConfigError:
        logger.info("EXCEL_MASTER_PATH nao configurada; ignorando sync de Empresas")
        return False


def append_filial(
    nome_empresa: str,
    id_empresa: str,
    nome_unidade: str,
    id_unidade: str,
    *,
    base_dir: Optional[Path] = None,
) -> bool:
    sheet_name = _sheet_name(FILIAIS_SHEET_ENV, DEFAULT_FILIAIS_SHEET)
    try:
        with _open_workbook() as (path, workbook):
            if sheet_name not in workbook.sheetnames:
                raise ExcelSyncError(f"Planilha '{path}' nao possui aba '{sheet_name}'")
            sheet = workbook[sheet_name]
            if _filial_exists(sheet, id_empresa, id_unidade):
                return False

            normalized_emp = _normalize_id(id_empresa, width=4)
            normalized_unit = _normalize_id(id_unidade, width=3)
            path_value = _path_str(base_dir, nome_empresa, normalized_emp, nome_unidade, normalized_unit)
            sheet.append([
                f"E-{normalized_emp}-F-{normalized_unit}",
                normalized_emp,
                normalized_unit,
                nome_empresa,
                nome_unidade,
                path_value,
            ])
            return True
    except ExcelSyncConfigError:
        logger.info("EXCEL_MASTER_PATH nao configurada; ignorando sync de Filiais")
        return False


def rename_empresa(id_empresa: str, novo_nome: str, *, base_dir: Optional[Path] = None) -> bool:
    sheet_emp = _sheet_name(EMPRESAS_SHEET_ENV, DEFAULT_EMPRESAS_SHEET)
    sheet_fil = _sheet_name(FILIAIS_SHEET_ENV, DEFAULT_FILIAIS_SHEET)
    updated = False
    normalized_emp = _normalize_id(id_empresa, width=4)

    with _open_workbook() as (path, workbook):
        if sheet_emp not in workbook.sheetnames:
            raise ExcelSyncError(f"Planilha '{path}' nao possui aba '{sheet_emp}'")
        empresas_ws = workbook[sheet_emp]
        for row in empresas_ws.iter_rows(min_row=2):
            if len(row) < 3:
                continue
            if _normalize_id(row[2].value, width=4) == normalized_emp:
                row[0].value = novo_nome
                if len(row) > 3:
                    row[3].value = novo_nome
                updated = True
                break

        if sheet_fil not in workbook.sheetnames:
            return updated

        filiais_ws = workbook[sheet_fil]
        for row in filiais_ws.iter_rows(min_row=2):
            if len(row) < 4:
                continue
            emp_id_cell = _normalize_id(row[1].value if len(row) > 1 else None, width=4)
            if emp_id_cell != normalized_emp:
                continue
            if len(row) > 3:
                row[3].value = novo_nome
            nome_unidade = row[4].value if len(row) > 4 and row[4].value else ""
            unidade_id = _normalize_id(row[2].value if len(row) > 2 else None, width=3)
            path = _path_str(base_dir, novo_nome, normalized_emp, nome_unidade, unidade_id)
            if len(row) > 5:
                row[5].value = path
            updated = True
    return updated


def rename_filial(
    id_empresa: str,
    id_unidade: str,
    novo_nome_unidade: str,
    *,
    nome_empresa: Optional[str] = None,
    base_dir: Optional[Path] = None,
) -> bool:
    sheet_fil = _sheet_name(FILIAIS_SHEET_ENV, DEFAULT_FILIAIS_SHEET)
    normalized_emp = _normalize_id(id_empresa, width=4)
    normalized_unit = _normalize_id(id_unidade, width=3)

    with _open_workbook() as (path, workbook):
        if sheet_fil not in workbook.sheetnames:
            raise ExcelSyncError(f"Planilha '{path}' nao possui aba '{sheet_fil}'")
        filiais_ws = workbook[sheet_fil]
        found = False
        for row_idx, row_values in enumerate(filiais_ws.iter_rows(min_row=2, values_only=True), start=2):
            emp_id = _normalize_id(row_values[1] if len(row_values) > 1 else None, width=4)
            unit_id = _normalize_id(row_values[2] if len(row_values) > 2 else None, width=3)
            if emp_id != normalized_emp or unit_id != normalized_unit:
                continue
            filiais_ws.cell(row=row_idx, column=5, value=novo_nome_unidade)
            if nome_empresa:
                filiais_ws.cell(row=row_idx, column=4, value=nome_empresa)
                empresa_nome = nome_empresa
            else:
                empresa_nome = row_values[3] if len(row_values) > 3 and row_values[3] else ""
            path_value = _path_str(base_dir, empresa_nome, normalized_emp, novo_nome_unidade, normalized_unit)
            filiais_ws.cell(row=row_idx, column=6, value=path_value)
            found = True
        return found




@dataclass
class EmpresaEmailRecord:
    """Empresas registradas na planilha com e-mails associados."""

    nome: str
    id_empresa: str
    emails: List[str]
    excel_rows: List[int]


_EMAIL_SPLIT_REGEX = re.compile(r"[;,\n]+")
_EMAIL_VALID_REGEX = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")


def _extract_emails(value: Optional[object]) -> List[str]:
    if value is None:
        return []
    text = str(value)
    if not text:
        return []

    parts = _EMAIL_SPLIT_REGEX.split(text)
    emails: List[str] = []
    seen = set()
    for raw in parts:
        candidate = raw.strip().strip('"').strip("'")
        if not candidate:
            continue
        candidate = candidate.replace(' ', '')
        if not _EMAIL_VALID_REGEX.match(candidate):
            continue
        lowered = candidate.lower()
        if lowered in seen:
            continue
        seen.add(lowered)
        emails.append(candidate)
    return emails


def fetch_empresas_emails(sheet_name: Optional[str] = None) -> Dict[str, EmpresaEmailRecord]:
    """Return a mapping id_empresa -> EmpresaEmailRecord using spreadsheet data."""

    chosen_sheet = sheet_name or _sheet_name(EMPRESAS_SHEET_ENV, DEFAULT_EMPRESAS_SHEET)
    path, workbook = _open_workbook_values()
    try:
        if chosen_sheet not in workbook.sheetnames:
            raise ExcelSyncError(f"Planilha '{path}' nao possui aba '{chosen_sheet}'")
        sheet = workbook[chosen_sheet]
        result: Dict[str, EmpresaEmailRecord] = {}
        for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row:
                continue
            nome = str(row[0]).strip() if len(row) > 0 and row[0] else ""
            id_empresa = _normalize_id(row[2] if len(row) > 2 else None, width=4)
            if not nome or not id_empresa.strip("0"):
                continue
            emails = _extract_emails(row[7] if len(row) > 7 else None)
            if not emails and id_empresa not in result:
                result[id_empresa] = EmpresaEmailRecord(nome=nome, id_empresa=id_empresa, emails=[], excel_rows=[idx])
                continue

            record = result.setdefault(
                id_empresa,
                EmpresaEmailRecord(nome=nome, id_empresa=id_empresa, emails=[], excel_rows=[])
            )
            record.excel_rows.append(idx)
            for email in emails:
                if email not in record.emails:
                    record.emails.append(email)
        return result
    finally:
        try:
            workbook.close()
        except Exception:
            pass
